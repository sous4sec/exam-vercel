from datetime import datetime
from pathlib import Path
import json
import copy

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from template_layout import LAYOUT


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _extrair_coluna_linha(celula):
    i = 0
    while i < len(celula) and celula[i].isalpha():
        i += 1
    return celula[:i], int(celula[i:])


def carregar_mapeamento_gabarito(caminho):
    with open(caminho, "r", encoding="utf-8") as f:
        return json.load(f)


def obter_carteiras_alunos_por_serie(numero_sala, mapeamento_gabarito):
    for sala_data in mapeamento_gabarito:
        if f"Sala {numero_sala}" == sala_data["sala"]:
            carteiras_por_serie = {}
            for posicao in sala_data["posicoes_mapeadas"]:
                if posicao["tipo"] == "CARTEIRA" and posicao["serie_obrigatoria"]:
                    serie = posicao["serie_obrigatoria"]
                    carteiras_por_serie.setdefault(serie, []).append({
                        "celula": posicao["celula"],
                        "serie": serie,
                        "cor": posicao["cor_obrigatoria"],
                    })
            return carteiras_por_serie
    return {}


def _safe_set(ws, col_letter, row, value):
    """Escreve na célula, respeitando mesclagens."""
    try:
        col_idx = column_index_from_string(col_letter)
    except ValueError:
        return

    for merged in ws.merged_cells.ranges:
        if (merged.min_row <= row <= merged.max_row
                and merged.min_col <= col_idx <= merged.max_col):
            ws.cell(row=merged.min_row, column=merged.min_col).value = value
            return

    ws[f"{col_letter}{row}"] = value


# ---------------------------------------------------------------------------
# preenchimento do mapa visual (carteiras coloridas)
# ---------------------------------------------------------------------------

def preencher_mapa_visual(ws, numero_sala, alunos, mapeamento_gabarito):
    carteiras_por_serie = obter_carteiras_alunos_por_serie(numero_sala, mapeamento_gabarito)
    if not carteiras_por_serie:
        return

    for serie_nome in ["1", "2", "3"]:
        lista_alunos = alunos.get(serie_nome, [])
        serie_completa = f"{serie_nome}º ano"
        if serie_completa not in carteiras_por_serie:
            continue
        carteiras = carteiras_por_serie[serie_completa]
        for indice, aluno in enumerate(lista_alunos):
            if indice < len(carteiras):
                coluna, linha = _extrair_coluna_linha(carteiras[indice]["celula"])
                _safe_set(ws, coluna, linha, aluno.nome)


# ---------------------------------------------------------------------------
# lista lateral coluna AA  — ordem alfabética (para consulta rápida)
# ---------------------------------------------------------------------------

def preencher_mapa(ws, sala_cfg, alunos):
    inicio = sala_cfg["inicio"]
    fim = sala_cfg["fim"]

    todos = sorted(
        alunos["1"] + alunos["2"] + alunos["3"],
        key=lambda a: a.nome.strip().lower(),
    )

    capacidade = fim - inicio + 1
    if len(todos) > capacidade:
        raise Exception(
            f"Sala suporta {capacidade} alunos e recebeu {len(todos)}"
        )

    for indice, aluno in enumerate(todos):
        _safe_set(ws, "AA", inicio + indice, aluno.nome)


# ---------------------------------------------------------------------------
# lista lateral coluna B  — ordem da distribuição aleatória (não alfabética)
# ---------------------------------------------------------------------------

def preencher_lista_lateral(ws, sala_cfg, alunos):
    """
    Preenche a lista lateral (coluna B) na ORDEM DA DISTRIBUIÇÃO — ou seja,
    a ordem em que os alunos foram sorteados para a sala. Isso garante que
    cada combinação/aba tenha uma lista diferente e coerente com o mapa visual.
    """
    listas = sala_cfg["lista"]

    for ano in ["1", "2", "3"]:
        linha_inicio = listas[ano]["inicio"]
        # SEM sorted() — mantém a ordem aleatória da distribuição
        for indice, aluno in enumerate(alunos[ano]):
            _safe_set(ws, "B", linha_inicio + indice, aluno.nome)


# ---------------------------------------------------------------------------
# cópia de aba (openpyxl não tem copy_worksheet entre workbooks nativamente)
# ---------------------------------------------------------------------------

def _copiar_aba(wb_origem, wb_destino, nome_origem, nome_destino):
    """Copia uma aba de wb_origem para wb_destino preservando valores e estilos básicos."""
    ws_orig = wb_origem[nome_origem]
    ws_dest = wb_destino.create_sheet(title=nome_destino)

    # Dimensões de colunas e linhas
    for col_letter, col_dim in ws_orig.column_dimensions.items():
        ws_dest.column_dimensions[col_letter].width = col_dim.width
        ws_dest.column_dimensions[col_letter].hidden = col_dim.hidden

    for row_num, row_dim in ws_orig.row_dimensions.items():
        ws_dest.row_dimensions[row_num].height = row_dim.height
        ws_dest.row_dimensions[row_num].hidden = row_dim.hidden

    # Mesclagens
    for merged in ws_orig.merged_cells.ranges:
        ws_dest.merge_cells(str(merged))

    # Células — valores e estilos
    for row in ws_orig.iter_rows():
        for cell in row:
            new_cell = ws_dest.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.border = copy.copy(cell.border)
                new_cell.alignment = copy.copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy.copy(cell.protection)

    return ws_dest


# ---------------------------------------------------------------------------
# função principal
# ---------------------------------------------------------------------------

def gerar_excel(template, output, lista_distribuicoes, caminho_mapeamento_gabarito=None):
    """
    Gera arquivo Excel com 6 abas, cada uma com uma combinação diferente.

    Args:
        template: Caminho do template
        output: Caminho do arquivo de saída
        lista_distribuicoes: Lista de 6 dicts {sala: {"1":[], "2":[], "3":[]}}
        caminho_mapeamento_gabarito: Caminho JSON de mapeamento (opcional)
    """
    # Carrega mapeamento de gabarito
    mapeamento_gabarito = None
    if caminho_mapeamento_gabarito and Path(caminho_mapeamento_gabarito).exists():
        try:
            mapeamento_gabarito = carregar_mapeamento_gabarito(caminho_mapeamento_gabarito)
        except Exception as e:
            print(f"Aviso: mapeamento não carregado: {e}")

    # Carrega template de referência (usado para copiar)
    wb_template = load_workbook(template)

    # Cria workbook de saída a partir do template (já tem a aba mod 1)
    wb_saida = load_workbook(template)

    # Remove aba Turmas se existir (item 3)
    for nome_remover in ["Turmas", "Planilha3"]:
        if nome_remover in wb_saida.sheetnames:
            del wb_saida[nome_remover]

    # Renomeia a aba existente "mod 1" para "Combinação 1"
    ws_existente = wb_saida["mod 1"]
    ws_existente.title = "Combinação 1"

    # Cria as demais 5 abas copiando do template
    for i in range(2, len(lista_distribuicoes) + 1):
        _copiar_aba(wb_template, wb_saida, "mod 1", f"Combinação {i}")

    # Preenche cada aba com sua distribuição
    for i, distribuicao in enumerate(lista_distribuicoes, start=1):
        ws = wb_saida[f"Combinação {i}"]

        for sala, alunos in distribuicao.items():
            cfg = LAYOUT[str(sala)]

            if mapeamento_gabarito:
                preencher_mapa_visual(ws, sala, alunos, mapeamento_gabarito)

            preencher_mapa(ws, cfg, alunos)
            preencher_lista_lateral(ws, cfg, alunos)

    # Salva
    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb_saida.save(output_path)
        return output_path
    except PermissionError as exc:
        fallback = output_path.with_name(
            f"{output_path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{output_path.suffix}"
        )
        wb_saida.save(fallback)
        return fallback
